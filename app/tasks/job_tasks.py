import io
import csv
import json
import asyncio
import os
import aioboto3
from datetime import datetime
from app.utils.celery_app import celery_app
from app.utils.dynamo import get_dynamodb, DYNAMODB_TABLE_NAME

CHANNEL = "notifications:job_events"


def _run_async(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()
        asyncio.set_event_loop(None)


def _get_session():
    import os
    from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
    engine = create_async_engine(
        os.getenv("DATABASE_URL", "postgresql+asyncpg://postgres:postgres@localhost:5432/operations_hub"),
        echo=False
    )
    return async_sessionmaker(bind=engine, class_=AsyncSession, expire_on_commit=False)


async def _publish_event(job_id: str, actor_id: str, state: str, subject: str, body: str):
    from app.utils.cache import redis_client
    event = json.dumps({
        "job_id": job_id,
        "actor_id": actor_id,
        "state": state,
        "subject": subject,
        "body": body,
    })
    await redis_client.publish(CHANNEL, event)

    session = get_dynamodb()
    async with session.resource("dynamodb") as dynamodb:
        table = await dynamodb.Table(DYNAMODB_TABLE_NAME)
        await table.put_item(Item={
            "request_id": job_id,
            "timestamp": datetime.utcnow().isoformat(),
            "actor_id": actor_id,
            "action_type": f"JOB_{state.upper()}",
            "module_source": "job-processor",
            "source_ip": "internal",
        })


@celery_app.task(bind=True, max_retries=3)
def process_csv_task(self, job_id: str, s3_key: str, actor_id: str):
    from app.repositories.job_repository import JobRepository
    from app.models.job import JobState

    async def _run():
        SessionLocal = _get_session()
        async with SessionLocal() as db:
            repo = JobRepository(db)
            job = await repo.get_by_id(job_id)
            job.state = JobState.running
            await repo.update(job)

        try:
            boto_session = aioboto3.Session(
                aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
                aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
                region_name=os.getenv("AWS_REGION"),
            )
            async with boto_session.client("s3") as s3:
                resp = await s3.get_object(Bucket=os.getenv("S3_BUCKET_NAME"), Key=s3_key)
                content = await resp["Body"].read()

            # Support both .csv and .xlsx
            if s3_key.endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                rows = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]
            else:
                rows = list(csv.DictReader(io.StringIO(content.decode("utf-8"))))

            result_key = s3_key.replace("csv/", "results/") + "_result.json"
            async with boto_session.client("s3") as s3:
                await s3.put_object(
                    Bucket=os.getenv("S3_BUCKET_NAME"),
                    Key=result_key,
                    Body=json.dumps(rows, default=str).encode(),
                    ContentType="application/json"
                )

            SessionLocal2 = _get_session()
            async with SessionLocal2() as db:
                repo = JobRepository(db)
                job = await repo.get_by_id(job_id)
                job.state = JobState.completed
                job.result_s3_key = result_key
                await repo.update(job)

            await _publish_event(job_id, actor_id, "completed",
                                 "CSV Import Completed",
                                 f"Your CSV job {job_id} completed. {len(rows)} rows processed.")
        except Exception as exc:
            if self.request.retries < self.max_retries:
                raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))
            SessionLocal3 = _get_session()
            async with SessionLocal3() as db:
                repo = JobRepository(db)
                job = await repo.get_by_id(job_id)
                job.state = JobState.failed
                await repo.update(job)
            await _publish_event(job_id, actor_id, "failed",
                                 "CSV Import Failed",
                                 f"Your CSV job {job_id} failed after all retries: {str(exc)}")

    _run_async(_run())


@celery_app.task(bind=True, max_retries=3)
def generate_report_task(self, job_id: str, actor_id: str, meta: dict):
    from app.repositories.job_repository import JobRepository
    from app.models.job import JobState

    async def _run():
        SessionLocal = _get_session()
        async with SessionLocal() as db:
            repo = JobRepository(db)
            job = await repo.get_by_id(job_id)
            job.state = JobState.running
            await repo.update(job)

        try:
            report = {"job_id": job_id, "actor_id": actor_id, "meta": meta, "status": "generated"}
            result_key = f"reports/{actor_id}/{job_id}.json"

            boto_session = aioboto3.Session(
                aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
                aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
                region_name=os.getenv("AWS_REGION"),
            )
            async with boto_session.client("s3") as s3:
                await s3.put_object(
                    Bucket=os.getenv("S3_BUCKET_NAME"),
                    Key=result_key,
                    Body=json.dumps(report).encode(),
                    ContentType="application/json"
                )
                url = await s3.generate_presigned_url(
                    "get_object",
                    Params={"Bucket": os.getenv("S3_BUCKET_NAME"), "Key": result_key},
                    ExpiresIn=3600,
                )

            SessionLocal2 = _get_session()
            async with SessionLocal2() as db:
                repo = JobRepository(db)
                job = await repo.get_by_id(job_id)
                job.state = JobState.completed
                job.result_s3_key = result_key
                await repo.update(job)

            await _publish_event(job_id, actor_id, "completed",
                                 "Report Ready",
                                 f"Your report {job_id} is ready. Download: {url}")
        except Exception as exc:
            if self.request.retries < self.max_retries:
                raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))
            SessionLocal3 = _get_session()
            async with SessionLocal3() as db:
                repo = JobRepository(db)
                job = await repo.get_by_id(job_id)
                job.state = JobState.failed
                await repo.update(job)
            await _publish_event(job_id, actor_id, "failed",
                                 "Report Generation Failed",
                                 f"Your report job {job_id} failed after all retries: {str(exc)}")

    _run_async(_run())
